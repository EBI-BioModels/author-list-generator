import csv
import openpyxl
import pandas as pd
import random
from unidecode import unidecode


FILE_NAME = "AuthorsListnContributions.xlsx"
wb = openpyxl.load_workbook(FILE_NAME)
# This workbook has only one sheet, therefore, we choose the active sheet directly
ws = wb.active


def combine_affiliations(affiliations: list[str]) -> str:
    s = ""
    for a in affiliations:
        if a is not None:
            s += a + "; "
    s.rstrip("; ")
    return s
def compute_affiliations_str(indexes: list[int]) -> str:
    org_indexes = [item for item in indexes if item > 0]
    org_indexes.sort()
    s_indexes = ""
    for index in org_indexes:
        s_indexes += str(index) + ", "
    s_indexes = s_indexes.rstrip(", ")
    return s_indexes


def generate_author_list():
    df = pd.DataFrame(ws.values)
    d_df = df.to_dict()
    first_name = d_df[0]
    del first_name[0]
    last_name = d_df[1]
    del last_name[0]
    email_address = d_df[2]
    del email_address[0]
    affiliation1 = d_df[3]
    del affiliation1[0]
    affiliation2 = d_df[4]
    del affiliation2[0]

    affiliation3 = d_df[5]
    del affiliation3[0]

    affiliation4 = d_df[6]
    del affiliation4[0]

    contribution = d_df[7]
    del contribution[0]

    authorship_order = d_df[8]
    del authorship_order[0]
    affiliation_list = {}
    affiliation_order = 0
    author_para = ""
    author_with_affiliation = {}
    for index in authorship_order:
        author_para += first_name[index] + " " + last_name[index]
        # print(item)
        # print(first_name[i] + "\t" + str(i) + "\t" + str(item[1]))
        # print(first_name[index] + " " + last_name[index])

        org1 = affiliation1[index]
        org2 = affiliation2[index]
        org3 = affiliation3[index]
        org4 = affiliation4[index]
        org1_index = -1
        org2_index = -1
        org3_index = -1
        org4_index = -1
        if org1 is not None:
            if org1 not in author_with_affiliation:
                author_with_affiliation[org1] = [index]
            else:
                author_with_affiliation[org1].append(index)
            org1_index = list(author_with_affiliation.keys()).index(org1) + 1

        if org2 is not None:
            if org2 not in author_with_affiliation:
                author_with_affiliation[org2] = [index]
            else:
                author_with_affiliation[org2].append(index)
            org2_index = list(author_with_affiliation.keys()).index(org2) + 1

        if org3 is not None:
            if org3 not in author_with_affiliation:
                author_with_affiliation[org3] = [index]
            else:
                author_with_affiliation[org3].append(index)
            org3_index = list(author_with_affiliation.keys()).index(org3) + 1

        if org4 is not None:
            if org4 not in author_with_affiliation:
                author_with_affiliation[org4] = [index]
            else:
                author_with_affiliation[org4].append(index)
            org4_index = list(author_with_affiliation.keys()).index(org4) + 1

        org_indexes = [org1_index, org2_index, org3_index, org4_index]
        affiliation_superscript = compute_affiliations_str(org_indexes)

        author_para += "<sup>" + affiliation_superscript + "</sup>"  # get_super(affiliation_superscript)
        author_para += ", "

    i = 1
    institutions = ""
    for item in author_with_affiliation:
        # print(str(i) + "\t" + str(author_with_affiliation[item]) + "\t" + item)
        # print(get_super(str(i)) + item)
        institutions += "<sup>" + str(i) + "</sup>" + item + "<br/>"
        i += 1

    file = open("output.html", "w")

    # Write HTML content
    file.write("<!DOCTYPE html><html>")
    file.write("<head>")
    file.write("<meta charset='utf-8'><title>Orders of the authors</title>")
    file.write("</head>")
    file.write("<body>")
    file.write("<h1>Orders of the authors</h1>")
    file.write("<p>" + author_para.rstrip(", ") + "</p>")
    file.write("<p>" + institutions + "</p>")
    file.write("</body>")
    file.write("</html>")

    file.close()

    print("The orders of the authors in your manuscript has been written in HTML file successfully.")


def customise_data():
    # as of writing this comment, A2:A84 -> Authors First name, B2:B84 --> Last name
    list = generate_random_sequence()
    df = pd.DataFrame(ws.values)
    ddf = df.to_dict()
    first_name = ddf[0]
    del first_name[0]
    last_name = ddf[1]
    del last_name[0]
    j = 2
    for i in list:
        print(i)
        print(first_name[i])
        print(last_name[i])
        ws["A" + str(j)] = first_name[i]
        ws["B" + str(j)] = last_name[i]
        j += 1
    wb.save(filename=FILE_NAME)


def print_rows():
    for row in ws.iter_rows(values_only=True):
        print(row)


def generate_random_sequence():
    # as of writing this comment, the sheet has A2:A84 -> Authors First name, B2:B84 --> Last name
    my_list = range(1, 84)
    random_sample = random.sample(my_list, 83)

    return random_sample


def generate_with_random_names():
    print_rows()
    generate_random_sequence()
    customise_data()
    generate_author_list()
    print_rows()


def generate_author_list_for_biorxiv():
    filename = "authors_list_biorxiv.tsv"
    # prepare_data()
    df = pd.DataFrame(ws.values)
    d_df = df.to_dict()
    with open(filename, mode='w', encoding='utf8', newline='\n') as tsv_file:
        tsv_writer = csv.writer(tsv_file, delimiter='\t', lineterminator='\n')
        tsv_writer.writerow(["Email", "Institution", "First Name", "Middle Name(s)/Initial(s)", "Last Name", "Suffix",
                             "Corresponding Author", "Home Page URL", "Collaborative Group/Consortium", "ORCiD"])
        for index, row in enumerate(ws.iter_rows(min_row=2)):
            first_name = row[0].value
            first_name = unidecode(first_name)
            last_name = row[1].value
            last_name = unidecode(last_name)
            email_address = row[2].value

            affiliation1 = row[3].value
            affiliation2 = row[4].value
            affiliation3 = row[5].value
            affiliation4 = row[6].value
            affiliation = combine_affiliations([affiliation1, affiliation2, affiliation3, affiliation4])
            affiliation = unidecode(affiliation)
            arr = [email_address, affiliation, first_name, "", last_name, "", "", "", "", ""]
            if email_address == "sheriff@ebi.ac.uk":
                arr = [email_address, affiliation, first_name, "", last_name, "", "X", "", "", ""]
            tsv_writer.writerow(arr)


if __name__ == "__main__":
    # generate_with_random_names()
    # generate_author_list()
    generate_author_list_for_biorxiv()
    # prepare_data()